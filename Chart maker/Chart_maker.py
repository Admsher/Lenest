import tkinter
from tkinter import ttk
from docxtpl import DocxTemplate 
import datetime
from tkinter import *
from openpyxl import load_workbook
from tkinter import messagebox
import os
import pandas as pd



Filename='C:/BITS/PS 1/New folder/Tests.xlsx'
excel_chart=load_workbook(Filename,data_only=True, keep_vba=True)
schedule_sheet=excel_chart[excel_chart.sheetnames[0]]
symptom_sheet=excel_chart[excel_chart.sheetnames[1]]
expected_date=1
Description=[None]*44
for i in range(0,41):    
 Description[i]=str(schedule_sheet.cell(row=(i+1),column=2).value)
weekwisearray=[None]*44
chart_list=[]
symptom_list=[None]*len(schedule_sheet['A'])
for i in range(0,len(schedule_sheet['A'])):    
 symptom_list[i]=str(symptom_sheet.cell(row=(i+2),column=1).value)





###Window 1 func.###

def chart_calculator():
      
        day_date=int(day.get())
        month_date=int(month.get())
        year_date=int(year.get())   
        expected_date= datetime.date(year_date,month_date,day_date)
        if(varPIH.get()=="True"):
            for i in range(0,41):    
                Description[i]=str(schedule_sheet.cell(row=(i+1),column=2).value)+","+str(schedule_sheet.cell(row=(i+2),column=3).value)
        if(varDiabetes.get()=="True"):
            for i in range(0,41):
                Description[i]=str(schedule_sheet.cell(row=(i+1),column=2).value)+","+str(schedule_sheet.cell(row=(i+2),column=4).value)
        if(varIUGI.get()=="True"):
            for i in range(0,41):    
                Description[i]=str(schedule_sheet.cell(row=(i+1),column=2).value)+","+str(schedule_sheet.cell(row=(i+2),column=5).value)
        if(varAnemia.get()=="True"):
            for i in range(0,41):
                Description[i]=str(schedule_sheet.cell(row=(i+1),column=2).value)+","+str(schedule_sheet.cell(row=(i+2),column=6).value)        
        for i in range (0,43):
            weekwisearray[i]=str(expected_date+datetime.timedelta(days=-7*(i)))
        for i in range(0,42):
              weektype=str(weekwisearray[42-i])+"-"+str(weekwisearray[41-i]) 
              Week="Week"+" "+str(i+1)
              procedure=Description[i]
              chart_tree=[weektype,procedure]
              chart_item=[Week,weektype,procedure]
              tree.insert('',0,values=chart_tree)
              chart_list.append(chart_item)
               
        
        messagebox.showinfo("Alert", "Calculation Complete")


def clear():
       
        day.delete(0,tkinter.END)
        month.delete(0,tkinter.END)
        year.delete(0,tkinter.END)
        day.insert(0,"1")
        month.insert(0,"1")
        year.insert(0,"2023")
        varPIH.set("False")
        varDiabetes.set("False")
        chart_list.clear() 
        tree.delete(*tree.get_children())
      


def confirm():
     doc=DocxTemplate('C:/BITS/PS 1/New folder/chart.docx')
     name=first_name_entry.get()+last_name_entry.get()
     day_date=int(day.get())
     month_date=int(month.get())
     year_date=int(year.get())   
     expected_date= datetime.date(year_date,month_date,day_date)
     doc.render({
          "name":name,
          "date":expected_date,
          "chart_list":chart_list,
     })

     doc_name=name+str(expected_date)+".docx"
     doc.save('C:/BITS/PS 1/New Folder/'+doc_name)
     messagebox.showinfo("Alert", "Document ready")
     os.startfile('C:/BITS/PS 1/New Folder/'+doc_name)


def Window2():
   ###Window 2 func.###  
     def result():
      
      symptom_prediction=[]
      Answersymptom1=tkinter.Label(frame2,text=SymptomVar1.get())
      Answersymptom1.grid(row=2,column=2)
      Answersymptom2=tkinter.Label(frame2,text=SymptomVar2.get())
      Answersymptom2.grid(row=3,column=2) 
      Answersymptom3=tkinter.Label(frame2,text=SymptomVar3.get())
      Answersymptom3.grid(row=4,column=2) 
      
    
      for i in range(2,5): 
        for j in range(3,9):
          if(SymptomVar1.get()==symptom_sheet.cell(row=i,column=j).value):
              symptom_prediction.append(symptom_sheet.cell(row=i,column=3).value)
              
              
                    
      for i in range(2,5): 
        for j in range(3,9):
          if(SymptomVar2.get()==symptom_sheet.cell(row=i,column=j).value):
              symptom_prediction.append(symptom_sheet.cell(row=i,column=3).value)
              
                    
      for i in range(2,5): 
        for j in range(3,9):
          if(SymptomVar3.get()==symptom_sheet.cell(row=i,column=j).value):
               symptom_prediction.append(symptom_sheet.cell(row=i,column=3).value)
      
      for i in range(2,5): 
          if(int(SystoleEntry.get())== int(symptom_sheet.cell(row=i,column=10).value)):
               symptom_prediction.append(symptom_sheet.cell(row=i,column=3).value)
      for i in range(2,5): 
          if(int(DiastoleLabelEntry.get())== int((symptom_sheet.cell(row=i,column=11).value))):
               symptom_prediction.append(symptom_sheet.cell(row=i,column=3).value)
      for i in range(2,5): 
          if(int(WeightEntry.get())== int((symptom_sheet.cell(row=i,column=12).value))):
               symptom_prediction.append(symptom_sheet.cell(row=i,column=3).value)                   
                    
      Result=tkinter.Label(frame2,text=max(set(symptom_prediction),key=symptom_prediction.count),font='bold')
      Result.grid(row=8,column=3)

   ###Window 2###
     window2=tkinter.Tk()
     window2.title("Predictor")
     frame2=tkinter.Frame(window2)
     frame2.pack(padx=20,pady=10)

     Symptom1Label=tkinter.Label(frame2,text="Symptom1")
     Symptom1Label.grid(row=2,column=0)
     SymptomVar1=StringVar()
     SymptomVar1.set("Select")
     Symptom1Entry=tkinter.OptionMenu(frame2,SymptomVar1,*symptom_list)
     Symptom1Entry.grid(row=2,column=1)
     
     

     Symptom2Label=tkinter.Label(frame2,text="Symptom2")
     Symptom2Label.grid(row=3,column=0)
     SymptomVar2=StringVar()
     SymptomVar2.set("Select")
     Symptom2Entry=tkinter.OptionMenu(frame2,SymptomVar2,*symptom_list)
     Symptom2Entry.grid(row=3,column=1)

     Symptom3Label=tkinter.Label(frame2,text="Symptom3")
     Symptom3Label.grid(row=4,column=0)
     SymptomVar3=StringVar()
     SymptomVar3.set("Select")
     Symptom3Entry=tkinter.OptionMenu(frame2,SymptomVar3,*symptom_list)
     Symptom3Entry.grid(row=4,column=1)
    
     
     
     SystoleLabel=tkinter.Label(frame2,text="Systole")
     SystoleLabel.grid(row=5,column=0)
     SystoleEntry=tkinter.Entry(frame2)
     SystoleEntry.grid(row=5,column=1)
     SystoleEntry.insert(0,"0")
     
     DiastoleLabel=tkinter.Label(frame2,text="Diastole")
     DiastoleLabel.grid(row=6,column=0)
     DiastoleLabelEntry=tkinter.Entry(frame2)
     DiastoleLabelEntry.grid(row=6,column=1)
     DiastoleLabelEntry.insert(0,"0")
     
     WeightIncrease=tkinter.Label(frame2,text="Weight Increase")
     WeightIncrease.grid(row=7,column=0)
     WeightEntry=tkinter.Entry(frame2)
     WeightEntry.grid(row=7,column=1)
     WeightEntry.insert(0,"0")
     
     Predict=tkinter.Button(frame2,text="Predict",command=result)
     Predict.grid(row=7,column=2)
     
     Answer=tkinter.Label(frame2,text="Predicted for:",font='bold')
     Answer.grid(row=8,column=2)

     window2.mainloop()
     



### Window 1####
window =tkinter.Tk()
window.title("Chart Generator")
frame1=tkinter.Frame(window)
style=ttk.Style(window)
style.theme_use('alt')
frame1.pack(padx=20,pady=10)


switch_button_1=tkinter.Button(frame1,text="<.>",command=Window2)
switch_button_1.grid(row=0,column=9)



first_name_label = tkinter.Label(frame1, text="First Name")
first_name_label.grid(row=1, column=0)
last_name_label = tkinter.Label(frame1, text="Last Name")
last_name_label.grid(row=1, column=1)


first_name_entry = tkinter.Entry(frame1)
last_name_entry = tkinter.Entry(frame1)
first_name_entry.grid(row=2, column=0)
last_name_entry.grid(row=2, column=1)

estimated_date_label=tkinter.Label(frame1,text="Estimated date")
estimated_date_label.grid(row=4,column=0)

day_12=tkinter.Label(frame1,text="Day:")
day_12.grid(row=5,column=0)
day=tkinter.Spinbox(frame1,from_=1,to=31)
day.grid(row=5,column=1)
month_12=tkinter.Label(frame1,text="Month:")
month_12.grid(row=6,column=0)
month=tkinter.Spinbox(frame1,from_=1,to=12)
month.grid(row=6,column=1)
year_12=tkinter.Label(frame1,text="Year:")
year_12.grid(row=7,column=0)
year=tkinter.Spinbox(frame1,from_=2023,to=2099)
year.grid(row=7,column=1)

complexities=tkinter.Label(frame1,text="Complexions")
complexities.grid(row=8,column=0)
varPIH=StringVar()
varPIH.set("False")
PIH=Checkbutton(frame1,text="PIH",variable=varPIH,onvalue="True",offvalue="False")
PIH.grid(row=8,column=1)
varDiabetes=StringVar()
varDiabetes.set("False")
Diabetes=Checkbutton(frame1,text="Diabetes",variable=varDiabetes,onvalue="True",offvalue="False")
Diabetes.grid(row=8,column=2)
varIUGI=StringVar()
varIUGI.set("False")
IUGI=Checkbutton(frame1,text="IUGI",variable=varIUGI,onvalue="True",offvalue="False")
IUGI.grid(row=8,column=3)
varAnemia=StringVar()
varAnemia.set("False")
Anemia=Checkbutton(frame1,text="Anemia",variable=varAnemia,onvalue="True",offvalue="False")
Anemia.grid(row=8,column=4)


Calculator=tkinter.Button(frame1,text="Calculate",command=chart_calculator)
Calculator.grid(row=4,column=2,columnspan=3,sticky="news",padx=20,pady=5)

Clear=tkinter.Button(frame1,text="Clear",command=clear)
Clear.grid(row=5,column=2,columnspan=3,sticky="news",padx=20,pady=5)

Clear=tkinter.Button(frame1,text="Confirm",command=confirm)
Clear.grid(row=6,column=2,columnspan=3,sticky="news",padx=20,pady=5)

columns=('Week No.','Description')
tree=ttk.Treeview(frame1,columns=columns,show="headings")
tree.heading('Week No.', text="Week no.")
tree.heading('Description',text="Procedures")
tree.grid(row=3,column=6,columnspan=3,padx=20,pady=10)



window.mainloop()