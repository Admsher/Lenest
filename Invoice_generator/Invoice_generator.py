import tkinter
from tkinter import ttk
from docxtpl import DocxTemplate 
import datetime
import time
from datetime import date
import pandas as pd
from tkinter import *
from openpyxl import load_workbook
from tkinter import messagebox
import os



Dest_filename = 'filepath'
df=pd.read_excel(Dest_filename) 
wb = load_workbook(Dest_filename)
ws=wb["Sheet1"]
df["medicine+price"]=df.MEDICINES.astype(str)+" -"+"₹"+df.PRICE.astype(str)
medprice=df["medicine+price"]
invoice_list=[]


def clear_item():
    qty_spinbox.delete(0,tkinter.END)
    qty_spinbox.insert(0, "1")
    unit_entry1.delete(0,tkinter.END)
    unit_entry1.insert(0,"0")
    


def add_item():
    qty=int(qty_spinbox.get())
    desc_temp=desc_entry.get(ANCHOR)
    desc=desc_temp[:-6]  
    price=float(unit_entry1.get())*(100-int(clicked1.get()))/100
    finale_price=price
    line_total=qty*(finale_price)
    invoice_item=[desc,qty,finale_price,line_total]
    tree.insert('',0,values=invoice_item)
    invoice_list.append(invoice_item)
    
    clear_item()

def new_invoice():
    first_name_entry.delete(0,tkinter.END)
    last_name_entry.delete(0,tkinter.END)
    phone_entry.delete(0,tkinter.END)
    clear_item()    
    tree.delete(*tree.get_children())
                
    invoice_list.clear()


def generate_invoice():
    doc=DocxTemplate("filepath")
    name=first_name_entry.get()+" "+last_name_entry.get()
    phone=phone_entry.get()
    subtotal=sum(item[3] for item in invoice_list)
    salestax=0.1
    total=subtotal*(1-salestax)
    invoice=int(ws['D2'].value)+1
    today=date.today()
    payment=clicked.get()
    doc.render({
        "date":today,
        "no":invoice,
        "name":name,
        "phone":phone,
        "invoice_list":invoice_list,
        "subtotal":subtotal,
        "total":total,
        "payment":payment 
        })
    doc_name=name+datetime.datetime.now().strftime("%Y-%m-%d-%H%M%S")+".docx"
    doc.save("folderpath"+doc_name)
    ws['D2'].value=invoice
    wb.save(Dest_filename)
    messagebox.showinfo("Invoice Complete", "Invoice Complete")
    os.startfile("folderpath"+doc_name)

window=tkinter.Tk()
window.title("Invoice Form")
frame=tkinter.Frame(window)
frame.pack(padx=20,pady=10)

first_name_label=tkinter.Label(frame,text="First Name")
first_name_label.grid(row=0,column=0)
first_name_entry=tkinter.Entry(frame)
first_name_entry.grid(row=1,column=0)

last_name_label=tkinter.Label(frame,text="Last Name")
last_name_entry=tkinter.Entry(frame)
last_name_label.grid(row=0,column=1)
last_name_entry.grid(row=1,column=1)

phone_label=tkinter.Label(frame,text="Phone")
phone_label.grid(row=0,column=2)
phone_entry=tkinter.Entry(frame)
phone_entry.grid(row=1,column=2)

qty_label=tkinter.Label(frame,text="Quantity")
qty_label.grid(row=2,column=0)
qty_spinbox=tkinter.Spinbox(frame,from_=1, to=100)
qty_spinbox.grid(row=3,column=0)

desc_entry=Listbox(window)
for item in medprice:
    desc_entry.insert(END,item)
desc_box=tkinter.Label(frame,text=desc_entry.get(ANCHOR))
desc_box.grid(row=6,column=1)
desc_entry.pack(padx=10,pady=10)

unit_price1_label= tkinter.Label(frame,text="Unit Price")
unit_price1_label.grid(row=2,column=1)
unit_entry1=tkinter.Spinbox(frame,from_=0,to=7000,increment=1)
unit_entry1.grid(row=3,column=1)

discount_label=tkinter.Label(frame,text="Discount")
discount_label.grid(row=2,column=2)
clicked1=StringVar()
clicked1.set("20")
discount_label_select=tkinter.OptionMenu(frame,clicked1,"0","20","30")
discount_label_select.grid(row=3,column=2)



clicked=StringVar()
clicked.set("Cash")
payment_method=tkinter.Label(frame,text="Payment mode")
payment_method.grid(row=2,column=3)
payment_select=tkinter.OptionMenu(frame,clicked,"Cash","Card","UPI","Online Banking")
payment_select.grid(row=3,column=3)



columns=('qty','desc','price','total')
tree=ttk.Treeview(frame,columns=columns,show="headings")
tree.heading('qty', text="Quantity")
tree.heading('desc',text="Description")
tree.heading('price',text="Unit Price")
tree.heading('total', text="Total")
tree.grid(row=5,column=0,columnspan=3,padx=20,pady=10)


add_item_button=tkinter.Button(frame,text="Add item",command= add_item)
save_invoice_button=tkinter.Button(frame,text="Generate Invoice",command=generate_invoice)
new_invoice_button=tkinter.Button(frame,text="New Invoice",command=new_invoice)
add_item_button.grid(row=4,column=2,pady=5)
save_invoice_button.grid(row=6,column=0,columnspan=3,sticky="news",padx=20,pady=5)
new_invoice_button.grid(row=7,column=0,columnspan=3,sticky="news",padx=20,pady=5)


window.mainloop()
