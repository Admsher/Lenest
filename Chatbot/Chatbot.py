

import streamlit as st
from streamlit_chat import message
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook


# scope=['https://www.googleapis.com/auth/spreadsheets','https://www.googleapis.com/auth/drive'] 
# creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
# gc = gspread.authorize(creds)

workbook=load_workbook("videos.xlsx",data_only=True) 
worksheet=workbook[workbook.sheetnames[0]]
if "history" not in st.session_state:
    st.session_state.history = []
if "key" not in st.session_state:
     st.session_state.key=[]    

def Hotstart(): 
    st.session_state.key.append(Hotstart)
    message("Hi this is a chatbot brought to you by Lenest,please select one of the option number to continue.",is_user=False)
    st.button("Weekly tips for Pregnancy.",on_click=Week_answer)
    st.button("Information Regarding Pregnancy and More",on_click=Pregnancy_Info)
    st.button("Other Issues",on_click=Other_Issues_Pre)


def Week_answer():
        st.session_state.key.append(Week_answer)    
        message_bot=("Please type in the Week No.(1-38)")
        st.text_input("Ask the bot here.Write stop to end the chat.", key="input_text2", on_change=weekChart)
        message(message_bot,is_user=False)
    
def Pregnancy_Info():    
        st.session_state.key.append(Pregnancy_Info)   
        message_bot=("Please select the topic you are interested in: ")
        message(message_bot,is_user=False)
        st.button("Post pregnancy.",on_click=Post_Pregnancy_options)
        st.button("Newborn Care",on_click=NewBorn_care)
        st.button("Diseases",on_click=Diseases)
        st.button("Tests and Vaccinations",on_click=Tests_and_vaccinations)
        st.button("Health and Lifestyle During Pregnancy",on_click=HealthEtLIfestyle_options)
def Other_Issues_Pre():  
        st.session_state.key.append("Other_Issues_Pre")     
        message_bot=("Please select one of the topics.")
        message(message_bot,is_user=False)
        st.button("For Men",on_click=For_Men)
        st.button("For Women",on_click=For_Women)
       
  
def weekChart():
     st.session_state.key.append(weekChart)   
     user_message = st.session_state.input_text2
     message(user_message,is_user=True)

     if user_message.isnumeric():
        message_bot=(worksheet.cell(int(user_message),15).value)+" - "+(worksheet.cell(int(user_message),16).value)    
        message(message_bot,is_user=False) 
        st.session_state.history.append(message_bot)     
     else:
        st.text_input("Ask the bot here.Write stop to end the chat.", key="input_text2", on_change=weekChart)
        message_bot=('Sorry I could not understand that.')
        message(message_bot,is_user=False) 
    

def Post_Pregnancy_options():
        st.session_state.key.append(Post_Pregnancy_options)   
        message_bot=("Please select on of the following")
        message(message_bot,is_user=False)
        st.button("Cesearean",on_click=Cesearian)
        st.button("Body changes and care",on_click=Body_Changes)
        st.button("Bleeding",on_click=Bleeding)
        st.button("Miscarriage and abortion",on_click=Miscarriage)
        st.button("Contraception",on_click=Contraception)


           
def NewBorn_care():
        st.session_state.key.append(NewBorn_care)   
        message("Here are some videos you asked for:")
      
        for i in range (0,(int(worksheet["AN3"].value))):
            if(worksheet.cell(int(i+2),30).value==None or worksheet.cell(int(i+2),31).value==None):
                continue
            message_bot=((worksheet.cell(int(i+2),30).value)+" - "+(worksheet.cell(int(i+2),31).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False)   
           
def Diseases():
        st.session_state.key.append(Diseases)   
        message_bot=("Please select the disease:")
        message(message_bot,is_user=False)
        st.button("Cervical Cancer",on_click=Cervical_Cancer)  
        st.button("Diabetes",on_click=Diabetes)
        st.button("Down's Syndrome",on_click=Downs_Syndrome)
        st.button("Thalassemia",on_click=Thalessemia)
        st.button("Thyroid",on_click=Thyroid)
        st.button("Vaginitis",on_click=Vaginitis)
        st.button("Blood Pressure",on_click=Blood_Pressure)
        st.button("Backache",on_click=Backache)
        st.button("Morning sickness",on_click=Morning_Sickness)
        st.button("Constipation",on_click=Constipation)
        
      
       
def Tests_and_vaccinations():
        st.session_state.key.append(Tests_and_vaccinations)   
        message("Here are some videos you asked for:")
        print(worksheet["AN2"].value)
        for i in range (0,int(worksheet["AN2"].value)+2):
            
            if(worksheet.cell(int(i+2),28).value==None or worksheet.cell(int(i+2),29).value==None):
                continue
            message_bot=((worksheet.cell(int(i+2),28).value)+" - "+(worksheet.cell(int(i+2),29).value))
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False)   
            st.session_state.history.append(message_bot)

def HealthEtLIfestyle_options():
        st.session_state.key.append(HealthEtLIfestyle_options)   
        message_bot=("Please select from the following:")  
        message(message_bot,is_user=False)
        st.button("Nutrition",on_click=Nutrition)
        st.button("Stress Management",on_click=Stress_Management)
        st.button("General Doubts",on_click=General_doubts)
        st.button("Relationship",on_click=Relationship)
        st.button("Pregnancy Issues",on_click=Pregnancy_Issues)
        st.button("Investigation of pregnancy",on_click=Investigation)

            


def For_Women():
            st.session_state.key.append(For_Women)   
            message("Here are some videos you asked for:")
            for i in range (0,(int(worksheet["AN5"].value)+1)):
                if(worksheet.cell(int(i+2),10).value==None or worksheet.cell(int(i+2),11).value==None):
                    continue
                message_bot=((worksheet.cell(int(i+2),10).value)+" - "+(worksheet.cell(int(i+2),11).value))
                if(message_bot==""):
                    message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
                message(message_bot,is_user=False)   
                st.session_state.history.append(message_bot)
def For_Men():
        st.session_state.key.append(For_Men)   
        message("Here are some videos you asked for:")
        for i in range (0,int(worksheet["AN4"].value+1)): 
            if(worksheet.cell(int(i+2),8).value==None or worksheet.cell(int(i+2),9).value==None):
                continue
            message_bot=((worksheet.cell(int(i+2),8).value)+" - "+(worksheet.cell(int(i+2),9).value))
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False)   
            st.session_state.history.append(message_bot)

def Cervical_Cancer():
        st.session_state.key.append(Cervical_Cancer)   
        for i in range(0,2):
            if(worksheet.cell(int(i+2),22).value==None or worksheet.cell(int(i+2),23).value==None):
                continue
            message_bot=((worksheet.cell(int(i+2),22).value)+" - "+(worksheet.cell(int(i+2),23).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False)

def Diabetes():
        st.session_state.key.append(Diabetes)   
        for i in range(0,2):
            if(worksheet.cell(int(i+4),22).value==None or worksheet.cell(int(i+4),23).value==None):
                continue
            message_bot=((worksheet.cell(int(i+4),22).value)+" - "+(worksheet.cell(int(i+4),23).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False) 

def Downs_Syndrome():
        st.session_state.key.append(Downs_Syndrome)   
        for i in range(0,2):
            if(worksheet.cell(int(i+6),22).value==None or worksheet.cell(int(i+6),23).value==None):
                continue
            message_bot=((worksheet.cell(int(i+6),22).value)+" - "+(worksheet.cell(int(i+6),23).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False)   

def Thalessemia():
        st.session_state.key.append(Thalessemia)   
        for i in range(0,3):
            if(worksheet.cell(int(i+10),22).value==None or worksheet.cell(int(i+10),23).value==None):
                continue
            message_bot=((worksheet.cell(int(i+10),22).value)+" - "+(worksheet.cell(int(i+10),23).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False) 

def Thyroid():
        st.session_state.key.append(Thyroid)   
        for i in range(0,2):
            if(worksheet.cell(int(i+12),22).value==None or worksheet.cell(int(i+12),23).value==None):
                continue
            message_bot=((worksheet.cell(int(i+12),22).value)+" - "+(worksheet.cell(int(i+12),23).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False) 
def Vaginitis():
        st.session_state.key.append(Vaginitis)   
        for i in range(0,2):
            if(worksheet.cell(int(i+15),22).value==None or worksheet.cell(int(i+15),23).value==None):
                continue
            message_bot=((worksheet.cell(int(i+15),22).value)+" - "+(worksheet.cell(int(i+15),23).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False)
def Blood_Pressure():
        st.session_state.key.append(Blood_Pressure)   
        for i in range(0,2):
            if(worksheet.cell(int(i+17),22).value==None or worksheet.cell(int(i+17),23).value==None):
                continue
            message_bot=((worksheet.cell(int(i+17),22).value)+" - "+(worksheet.cell(int(i+17),23).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False) 
def Backache():
        st.session_state.key.append(Backache)   
        for i in range(0,2):
            if(worksheet.cell(int(i+19),22).value==None or worksheet.cell(int(i+19),23).value==None):
                continue
            message_bot=((worksheet.cell(int(i+19),22).value)+" - "+(worksheet.cell(int(i+19),23).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False) 
def Morning_Sickness():
        st.session_state.key.append(Morning_Sickness)   
        for i in range(0,2):
            if(worksheet.cell(int(i+21),22).value==None or worksheet.cell(int(i+21),23).value==None):
                continue
            message_bot=((worksheet.cell(int(i+21),22).value)+" - "+(worksheet.cell(int(i+21),23).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False)  
def Constipation():
        st.session_state.key.append(Constipation)   
        for i in range(0,2):
            if(worksheet.cell(int(i+23),22).value==None or worksheet.cell(int(i+23),23).value==None):
                continue
            message_bot=((worksheet.cell(int(i+23),22).value)+" - "+(worksheet.cell(int(i+23),23).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False)   
def Thyroid():
        st.session_state.key.append(Thyroid)   
        for i in range(0,2):
            if(worksheet.cell(int(i+26),22).value==None or worksheet.cell(int(i+26),23).value==None):
                continue
            message_bot=((worksheet.cell(int(i+26),22).value)+" - "+(worksheet.cell(int(i+26),23).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False)                                                


def Cesearian():
        st.session_state.key.append(Cesearian)   
        for i in range(0,2):
            if(worksheet.cell(int(i+2),25).value==None or worksheet.cell(int(i+2),26).value==None):
                continue
            message_bot=((worksheet.cell(int(i+2),25).value)+" - "+(worksheet.cell(int(i+2),26).value))
            st.session_state.history.append(message_bot)
            message(message_bot,is_user=False) 
def Body_Changes():
        st.session_state.key.append(Body_Changes)   
        for i in range(0,7):
            if(worksheet.cell(int(i+4),25).value==None or worksheet.cell(int(i+4),26).value==None):
                continue
            message_bot=((worksheet.cell(int(i+4),25).value)+" - "+(worksheet.cell(int(i+4),26).value))
            st.session_state.history.append(message_bot)
            message(message_bot,is_user=False) 

def Bleeding():
        st.session_state.key.append(Bleeding)   
        for i in range(0,2):
            if(worksheet.cell(int(i+11),25).value==None or worksheet.cell(int(i+11),26).value==None):
                continue
            message_bot=((worksheet.cell(int(i+11),25).value)+" - "+(worksheet.cell(int(i+11),26).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False)   

def Miscarriage():
        st.session_state.key.append(Miscarriage)   
        for i in range(0,3):
            if(worksheet.cell(int(i+14),25).value==None or worksheet.cell(int(i+14),26).value==None):
                continue
            message_bot=((worksheet.cell(int(i+14),25).value)+" - "+(worksheet.cell(int(i+14),26).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False) 

def Contraception():
        st.session_state.key.append(Contraception)   
        for i in range(0,2):
            if(worksheet.cell(int(i+17),25).value==None or worksheet.cell(int(i+17),26).value==None):
                continue
            message_bot=((worksheet.cell(int(i+17),25).value)+" - "+(worksheet.cell(int(i+17),26).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False) 

def Nutrition():
        st.session_state.key.append(Nutrition)   
        for i in range(0,10):
            if(worksheet.cell(int(i+2),32).value==None or worksheet.cell(int(i+2),33).value==None):
                continue
            message_bot=((worksheet.cell(int(i+2),32).value)+" - "+(worksheet.cell(int(i+2),33).value))
            st.session_state.history.append(message_bot)
            message(message_bot,is_user=False)
def Stress_Management():
        st.session_state.key.append(Stress_Management)   
        for i in range(0,10):
            if(worksheet.cell(int(i+13),32).value==None or worksheet.cell(int(i+13),33).value==None):
                continue
            message_bot=((worksheet.cell(int(i+13),32).value)+" - "+(worksheet.cell(int(i+13),33).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False) 
def General_doubts():
        st.session_state.key.append(General_doubts)   
        for i in range(0,9):
            if(worksheet.cell(int(i+24),32).value==None or worksheet.cell(int(i+24),33).value==None):
                continue
            message_bot=((worksheet.cell(int(i+24),32).value)+" - "+(worksheet.cell(int(i+24),33).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False)   
def Relationship():
        st.session_state.key.append(Relationship)   
        for i in range(0,5):
            if(worksheet.cell(int(i+34),32).value==None or worksheet.cell(int(i+34),33).value==None):
                continue
            message_bot=((worksheet.cell(int(i+34),32).value)+" - "+(worksheet.cell(int(i+34),33).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False) 
def Pregnancy_Issues():
        st.session_state.key.append(Pregnancy_Issues)   
        for i in range(0,7):
            if(worksheet.cell(int(i+40),32).value==None or worksheet.cell(int(i+40),33).value==None):
                continue
            message_bot=((worksheet.cell(int(i+40),32).value)+" - "+(worksheet.cell(int(i+40),33).value))
            st.session_state.history.append(message_bot)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
            message(message_bot,is_user=False) 
def Investigation():
        st.session_state.key.append(Investigation)   
        for i in range(0,8):
            if(worksheet.cell(int(i+48),32).value==None or worksheet.cell(int(i+48),33).value==None):
                continue
            message_bot=((worksheet.cell(int(i+48),32).value)+" - "+(worksheet.cell(int(i+48),33).value))
            st.session_state.history.append(message_bot)
            message(message_bot,is_user=False)
            if(message_bot==""):
                message_bot("Sorry I could not find anything on that, you might need to contact doctor for consultation.Sorry for the inconvenience.") 
    
def Goback():
     if len(st.session_state.key)==0 or len(st.session_state.key)==1:
          None
     else:
        del st.session_state.key[-1]
        st.session_state.key[len(st.session_state.key)-1]()
          
       
          
###Start Sidebar###
st.sidebar.markdown("""
<style>
.big-font {
    font-size:50px !important;
}
</style>
""", unsafe_allow_html=True)
st.sidebar.markdown('<p class="big-font">Le-Nest Chatbot</p>', unsafe_allow_html=True)
st.sidebar.write("This bot provides you with suggestions from the vast amount of videos Le NEST has curated for their patients based on your query.Please note that this is an automated service and does not mean to provide you with wrong or hurtful information.We would love to hear your review on this. ")
st.sidebar.button('Initiate/Restart chat',on_click=Hotstart)
st.sidebar.button('Go back',on_click=Goback)
c=st.sidebar.container()
c.write("See your previous answers here:")
for repeat in range(0,len(st.session_state.history)):
            c.write(st.session_state.history[len(st.session_state.history)-repeat-1])  






